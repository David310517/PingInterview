"""
Circuit Info Collector201
======================

A PySide6-based GUI tool for network engineers to **collect and organize WAN circuit and interface information from Cisco IOS devices**.

Features:
---------
- **User-Friendly GUI:**  
  - Select input Excel file (with device IP, hostname, site name).
  - Select output folder for reports.
  - Enter SSH credentials (username, password, enable secret).
  - Multi-select sites to process.
  - Preview all devices by site before running.
  - Progress bar and status for each device.

- **Excel Device List Parsing:**  
  - Auto-detects columns for IP, hostname, and site name using fuzzy matching.

- **Automated Network Collection (per device):**  
  - SSH (Netmiko) to Cisco IOS device.
  - Caches device config locally for efficiency.
  - Runs and extracts:
    - `show run` (full config, circuit-related interfaces, tunnels, BDIs, VLANs).
    - `show vrf` (VRF summary).
    - `show ip arp vrf <vrf>` (per VRF ARP tables).
    - `show ip vrf <vrf>` (VRF config details).
    - `show cdp neighbor detail`.
    - All referenced QoS policy-map blocks.

- **Data Extraction/Formatting:**  
  - Detects circuit blocks (tunnels, keyword matches, BDIs, VLANs) in running config.
  - Extracts all referenced policy-map details.
  - Gathers VRF and ARP details.
  - Writes per-device, per-site results to Excel, formatted and grouped.
  - Autosizes Excel columns for readability.

- **Failure Tracking:**  
  - Writes unreachable devices with error reasons to a `.txt` file.

- **Persistent Config:**  
  - Remembers last-used file paths and credentials between runs.

- **Modern Look:**  
  - Uses qdarkstyle for a professional dark theme.

Dependencies:
-------------
- Python 3.x
- PySide6
- pandas
- openpyxl
- netmiko
- qdarkstyle

How to Use:
-----------
1. Launch the script: `python circuitinfo20.py`
2. Select Excel input and output folder.
3. Enter SSH credentials.
4. Select one or more sites (multi-select list).
5. Optionally preview devices.
6. Click "Start". Progress and errors will be shown in the GUI.
7. Output Excel and a log of unreachable devices will be saved in your output folder.

Typical Use Cases:
------------------
- Auditing SD-WAN, MPLS, and WAN circuits before migration.
- Inventory collection for consulting/network projects.
- Quick "bulk" config collection for multi-site cutovers

"""
import os
import re
import configparser
import pandas as pd
from datetime import datetime
from netmiko import ConnectHandler
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from PySide6.QtCore import QObject, Signal, QThread, Qt
from PySide6.QtGui import QStandardItemModel, QStandardItem
from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QFormLayout, QHBoxLayout,
    QVBoxLayout, QLineEdit, QPushButton, QFileDialog, QProgressBar,
    QTableView, QMessageBox, QListWidget, QListWidgetItem, QLabel
)
import qdarkstyle

# ========== CONFIGURATION AND CONSTANTS ==========
CFG_FILE = os.path.expanduser("~/.circuit_info_config.ini")
KEYWORDS = ["ewans", "owans", "cid#", "rcn", "crowncastle", "comcast", "wans", "cid"]
TUNNEL_RE = re.compile(r"^interface\s+Tunnel", re.IGNORECASE)
BD_RE = re.compile(r"bridge-domain\s+(\d+)", re.IGNORECASE)
SITE_FILL = PatternFill("solid", fgColor="FFFF00")
SITE_FONT = Font(bold=True)
SITE_ALIGN = Alignment(horizontal="center")
HOST_FONT = Font(color="0000FF", underline="single")
CONFIG_FONT = Font(name="Consolas", size=10)

def load_cfg():
    """Load configuration (paths, credentials) from the user's home directory."""
    cfg = configparser.ConfigParser()
    if os.path.exists(CFG_FILE):
        cfg.read(CFG_FILE)
    return cfg

def save_cfg(cfg):
    """Save configuration to file for next launch."""
    with open(CFG_FILE, "w") as f:
        cfg.write(f)

def normalize_column_name(name: str) -> str:
    """Normalize DataFrame column names for fuzzy/alias matching."""
    return re.sub(r"[^a-z]", "", name.lower())

def find_column(df: pd.DataFrame, aliases):
    """
    Try to find a column in DataFrame using a list of possible aliases.
    Returns the column name as present in the DataFrame, or None.
    """
    norm = {normalize_column_name(c): c for c in df.columns}
    for a in aliases:
        na = normalize_column_name(a)
        if na in norm:
            return norm[na]
    return None

def extract_circuits(lines):
    """
    Given the lines of a 'show run' config, extract circuit-related interface blocks.
    """
    blocks, cur = [], []
    for l in lines:
        if l.startswith("interface"):
            if cur: blocks.append(cur)
            cur = [l.strip()]
        elif cur and l.startswith(" "):
            cur.append(l.strip())
        else:
            if cur: blocks.append(cur); cur = []
    if cur: blocks.append(cur)

    idx = {b[0]: b for b in blocks}
    sel, seen = [], set()
    for b in blocks:
        txt = " ".join(b).lower()
        if TUNNEL_RE.match(b[0].lower()) or any(kw in txt for kw in KEYWORDS):
            sel.append(b); seen.add(b[0])
    for b in sel[:]:
        for l in b:
            m = BD_RE.search(l)
            if m:
                k = f"interface BDI{m.group(1)}"
                if k in idx and k not in seen:
                    sel.append(idx[k]); seen.add(k)
            if "switchport trunk allowed vlan" in l:
                for v in re.findall(r"\d+", l):
                    k = f"interface Vlan{v}"
                    if k in idx and k not in seen:
                        sel.append(idx[k]); seen.add(k)
    return sel

def extract_policy_map(name, cfg):
    """
    Extract a 'policy-map <name>' block from running config (as a list of lines).
    Returns all lines in the block (indented).
    """
    blk, cap = [], False
    for l in cfg:
        if cap:
            if not l.startswith(" "): break
            blk.append(l.strip())
        elif l.strip().lower().startswith(f"policy-map {name.lower()}"):
            blk.append(l.strip()); cap = True
    return blk

def fetch_all(ip, name, user, pwd, sec, site_dir, site_key):
    """
    Connects to the device and retrieves:
    - 'show run', 'show vrf', 'show ip arp vrf', 'show ip vrf', QoS policy-maps, CDP neighbors.
    - Caches 'show run' to file for efficiency.
    Returns all outputs in a tuple. In case of error, returns error message in last field.
    """
    try:
        conn = ConnectHandler(
            device_type='cisco_ios', ip=ip,
            username=user, password=pwd, secret=sec, fast_cli=False
        )
        conn.enable()
        conn.send_command("terminal length 0")

        txt_file = os.path.join(site_dir, f"{name}_{ip}_showrun.txt")
        if os.path.exists(txt_file):
            with open(txt_file, "r") as f:
                cfg = [l.rstrip() for l in f.readlines()]
        else:
            cfg = conn.send_command("show run").splitlines()
            with open(txt_file, "w") as f:
                f.write("\n".join(cfg))

        raw_vrf = conn.send_command("show vrf").splitlines()
        vrf_names = []
        for l in raw_vrf:
            l = l.strip()
            if l and not l.lower().startswith("name"):
                parts = l.split()
                if parts:
                    vrf_names.append(parts[0])

        vrf_out = raw_vrf
        vrf_details, arp = {}, {}
        for v in vrf_names:
            arp[v] = conn.send_command(f"show ip arp vrf {v}").splitlines()
            vrf_details[v] = conn.send_command(f"show ip vrf {v}").splitlines()

        qos_refs = re.findall(r"policy-map\s+(qos-[\w-]+)", "\n".join(cfg), re.IGNORECASE)
        qos_blks = []
        for p in qos_refs:
            pm = extract_policy_map(p, cfg)
            if pm: qos_blks.extend(pm + [""])

        cdp_neighbor_detail = conn.send_command("show cdp neighbor detail").splitlines()
        conn.disconnect()
        return name, ip, cfg, vrf_out, vrf_details, arp, qos_blks, cdp_neighbor_detail, ""
    except Exception as e:
        return name, ip, [], [], {}, {}, [], [], str(e)

class Worker(QObject):
    device_done = Signal(int, str, str, bool, str)
    progress    = Signal(int)
    finished    = Signal(str)
    error       = Signal(str)

    def __init__(self, excel, output, user, pwd, sec, selected_sites):
        super().__init__()
        self.excel, self.output_dir = excel, output
        self.user, self.pwd, self.sec = user, pwd, sec
        self.selected_sites = selected_sites

    def run(self):
        try:
            df = pd.read_excel(self.excel)
            ip_col   = find_column(df, ["ip","ipaddress"])
            site_col = find_column(df, ["sitename","siteid"])
            host_col = find_column(df, ["hostname","switch"])

            df = df[df[site_col].isin(self.selected_sites)]

            devices = []
            for site, grp in df.groupby(site_col):
                safe = re.sub(r'[^0-9A-Za-z_.-]',"_",str(site))[:31] or "Site"
                site_dir = os.path.join(self.output_dir, safe)
                os.makedirs(site_dir, exist_ok=True)
                for _, r in grp.iterrows():
                    devices.append((safe, str(r[host_col]), str(r[ip_col]), site_dir))

            total = len(devices)
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            base = os.path.join(self.output_dir, f"Circuit_Info_output_{ts}")
            os.makedirs(base, exist_ok=True)
            wb = Workbook(); wb.remove(wb.active)

            site_next_col = {}
            failed_devices = []

            for idx, (site, host, ip, site_dir) in enumerate(devices, start=1):
                if site not in wb.sheetnames:
                    ws = wb.create_sheet(site)
                    site_next_col[site] = 1
                else:
                    ws = wb[site]
                    site_next_col[site] += 1

                dev_col = site_next_col[site]
                ws.cell(1, dev_col, site).fill = SITE_FILL
                ws.cell(1, dev_col).font = SITE_FONT
                ws.cell(1, dev_col).alignment = SITE_ALIGN

                row = 2
                ws.cell(row, dev_col, host).font = HOST_FONT
                ws.cell(row+1, dev_col, ip).font = HOST_FONT
                row += 3

                name, ip2, cfg, vrf_o, vrf_d, arp_d, qos, cdp_neighbor_detail, err = fetch_all(
                    ip, host, self.user, self.pwd, self.sec, site_dir, site
                )

                if err:
                    failed_devices.append((site, host, ip, err))

                interface_blocks = extract_circuits(cfg)
                if interface_blocks:
                    for blk in interface_blocks:
                        for l in blk:
                            c = ws.cell(row, dev_col, l)
                            c.font = CONFIG_FONT
                            c.alignment = Alignment(wrap_text=True)
                            row += 1
                        row += 1
                else:
                    ws.cell(row, dev_col, "No circuit interfaces detected on this device.").font = CONFIG_FONT
                    row += 2

                if vrf_o:
                    ws.cell(row, dev_col, "! --- show vrf summary ---").font = CONFIG_FONT
                    row += 1
                    for l in vrf_o:
                        ws.cell(row, dev_col, l).font = CONFIG_FONT
                        row += 1
                    row += 1

                if arp_d:
                    for v in arp_d:
                        ws.cell(row, dev_col, f"! show ip arp vrf {v}").font = CONFIG_FONT
                        row += 1
                        for l in arp_d[v]:
                            ws.cell(row, dev_col, l).font = CONFIG_FONT
                            row += 1

                        ws.cell(row, dev_col, f"! show ip vrf {v}").font = CONFIG_FONT
                        row += 1
                        for l in vrf_d.get(v, []):
                            ws.cell(row, dev_col, l).font = CONFIG_FONT
                            row += 1
                        row += 1

                if qos:
                    ws.cell(row, dev_col, "! --- QoS policy-map blocks ---").font = CONFIG_FONT
                    row += 1
                    for l in qos:
                        ws.cell(row, dev_col, l).font = CONFIG_FONT
                        row += 1
                    row += 1

                if cdp_neighbor_detail:
                    ws.cell(row, dev_col, "! --- show cdp neighbor detail ---").font = CONFIG_FONT
                    row += 1
                    for l in cdp_neighbor_detail:
                        ws.cell(row, dev_col, l).font = CONFIG_FONT
                        row += 1
                    row += 1

                col_cells = ws[get_column_letter(dev_col)]
                width = max((len(str(c.value)) for c in col_cells if c.value), default=0) + 2
                ws.column_dimensions[get_column_letter(dev_col)].width = min(width, 50)

                self.device_done.emit(idx-1, host, ip, err=="", err)
                self.progress.emit(int(idx/total * 100))

            out = os.path.join(base, f"Circuit_Info_{ts}.xlsx")
            for ws in wb.worksheets:
                ws.cell(1, 1).fill = SITE_FILL
                ws.cell(1, 1).font = SITE_FONT
                ws.cell(1, 1).alignment = SITE_ALIGN

            wb.save(out)
            fail_txt = os.path.join(self.output_dir, f"unreachable_devices_{ts}.txt")
            with open(fail_txt, "w") as f:
                if failed_devices:
                    for site, host, ip, err in failed_devices:
                        f.write(f"Site: {site}\tHost: {host}\tIP: {ip}\tReason: {err}\n")
                else:
                    f.write("All devices were reachable.\n")

            self.finished.emit(out)

        except Exception:
            import traceback
            self.error.emit(traceback.format_exc())

class MainWindow(QMainWindow):
    """
    Main PySide6 GUI window. 
    - File selectors for Excel and output.
    - Credentials entry.
    - Site selection (multi-select).
    - Preview of selected devices.
    - Progress bar and status table.
    - Threaded worker for device polling.
    """
    def __init__(self):
        super().__init__()
        self.setWindowTitle("📡 Circuit Info Collector")
        self.resize(950,650)
        cfg = load_cfg()

        cw = QWidget(self); self.setCentralWidget(cw)
        main = QVBoxLayout(cw); form = QFormLayout(); main.addLayout(form)

        self.excel_in = QLineEdit(cfg.get("DEFAULT","excel_path",fallback=""))
        btn_e = QPushButton("Browse…"); btn_e.clicked.connect(self.browse_excel)
        row1 = QHBoxLayout(); row1.addWidget(self.excel_in); row1.addWidget(btn_e)
        form.addRow("Excel File:", row1)

        self.out_in = QLineEdit(cfg.get("DEFAULT","output_dir",fallback=""))
        btn_o = QPushButton("Browse…"); btn_o.clicked.connect(self.browse_output)
        row2 = QHBoxLayout(); row2.addWidget(self.out_in); row2.addWidget(btn_o)
        form.addRow("Output Folder:", row2)

        self.user = QLineEdit(cfg.get("DEFAULT","ssh_username",fallback=""))
        form.addRow("SSH User:", self.user)
        self.pwd  = QLineEdit(cfg.get("DEFAULT","ssh_password",fallback=""))
        self.pwd.setEchoMode(QLineEdit.Password); form.addRow("SSH Pass:", self.pwd)
        self.sec  = QLineEdit(cfg.get("DEFAULT","enable_secret",fallback=""))
        self.sec.setEchoMode(QLineEdit.Password); form.addRow("Enable Sec:", self.sec)

        # Enhanced: Site selection widget
        self.site_list_label = QLabel("Select Sites:")
        self.site_list = QListWidget()
        self.site_list.setSelectionMode(QListWidget.MultiSelection)
        form.addRow(self.site_list_label, self.site_list)

        # Add Select All and Deselect All buttons
        select_buttons = QHBoxLayout()
        self.select_all_btn = QPushButton("Select All")
        self.select_all_btn.clicked.connect(self.select_all_sites)
        self.deselect_all_btn = QPushButton("Deselect All")
        self.deselect_all_btn.clicked.connect(self.deselect_all_sites)
        select_buttons.addWidget(self.select_all_btn)
        select_buttons.addWidget(self.deselect_all_btn)
        form.addRow(select_buttons)

        # Live Preview of selected devices (optional)
        self.preview_label = QLabel("Preview Devices in Selected Sites:")
        self.preview_model = QStandardItemModel(0, 3, self)
        self.preview_model.setHorizontalHeaderLabels(["Site", "Host", "IP"])
        self.preview_table = QTableView()
        self.preview_table.setModel(self.preview_model)
        self.preview_table.horizontalHeader().setStretchLastSection(True)
        main.addWidget(self.preview_label)
        main.addWidget(self.preview_table)

        btns = QHBoxLayout()
        self.start_btn = QPushButton("Start"); self.start_btn.clicked.connect(self.start)
        self.exit_btn  = QPushButton("Exit");  self.exit_btn.clicked.connect(self.close)
        btns.addWidget(self.start_btn); btns.addWidget(self.exit_btn)
        form.addRow(btns)

        self.pbar = QProgressBar(); main.addWidget(self.pbar)
        self.model = QStandardItemModel(0,5,self)
        self.model.setHorizontalHeaderLabels(["#","Host","IP","OK","Reason"])
        self.table = QTableView(); self.table.setModel(self.model)
        self.table.horizontalHeader().setStretchLastSection(True)
        main.addWidget(self.table)

        self.excel_in.textChanged.connect(self.populate_sites)
        self.site_list.itemSelectionChanged.connect(self.preview_devices)

        if self.excel_in.text().strip():
            self.populate_sites()

    def select_all_sites(self):
        """Check all sites in the site selection widget."""
        for i in range(self.site_list.count()):
            item = self.site_list.item(i)
            item.setCheckState(Qt.Checked)

    def deselect_all_sites(self):
        """Uncheck all sites in the site selection widget."""
        for i in range(self.site_list.count()):
            item = self.site_list.item(i)
            item.setCheckState(Qt.Unchecked)

    def browse_excel(self):
        p,_ = QFileDialog.getOpenFileName(self,"Select Excel",filter="Excel Files (*.xlsx)")
        if p: self.excel_in.setText(p)

    def browse_output(self):
        p = QFileDialog.getExistingDirectory(self,"Select Output Folder")
        if p: self.out_in.setText(p)

    def populate_sites(self):
        """Populate the site list only when required."""
        self.site_list.clear()
        if not self.excel_in.text().strip():
            return
        
        try:
            df = pd.read_excel(self.excel_in.text())
            site_col = find_column(df, ["sitename", "siteid"])
            sites = sorted(set(df[site_col].dropna()))
            for site in sites:
                item = QListWidgetItem(str(site))
                item.setFlags(item.flags() | Qt.ItemIsUserCheckable)
                item.setCheckState(Qt.Unchecked)
                self.site_list.addItem(item)
        except Exception as e:
            print(f"Error populating sites: {e}")
        self.preview_devices()

    def preview_devices(self):
        self.preview_model.removeRows(0, self.preview_model.rowCount())
        if not self.excel_in.text().strip():
            return

        try:
            df = pd.read_excel(self.excel_in.text())
            site_col = find_column(df, ["sitename", "siteid"])
            host_col = find_column(df, ["hostname", "switch"])
            ip_col = find_column(df, ["ip", "ipaddress"])
            selected_sites = [self.site_list.item(i).text() 
                              for i in range(self.site_list.count()) 
                              if self.site_list.item(i).checkState() == Qt.Checked]
            df = df[df[site_col].isin(selected_sites)]
            for _, row in df.iterrows():
                r = self.preview_model.rowCount()
                self.preview_model.insertRow(r, [
                    QStandardItem(str(row[site_col])),
                    QStandardItem(str(row[host_col])),
                    QStandardItem(str(row[ip_col]))
                ])
        except Exception as e:
            print(f"Error previewing devices: {e}")

    def get_selected_sites(self):
        return [self.site_list.item(i).text() 
                for i in range(self.site_list.count()) 
                if self.site_list.item(i).checkState() == Qt.Checked]

    def start(self):
        if not all([
            self.excel_in.text().strip(),
            self.out_in.text().strip(),
            self.user.text().strip(),
            self.pwd.text().strip(),
            self.sec.text().strip()
        ]):
            QMessageBox.warning(self,"Missing Data","Please fill all fields.")
            return

        selected_sites = self.get_selected_sites()
        if not selected_sites:
            QMessageBox.warning(self,"No Sites Selected","Please select at least one site to process.")
            return

        cfg = configparser.ConfigParser()
        cfg['DEFAULT'] = {
            'excel_path':    self.excel_in.text(),
            'output_dir':    self.out_in.text(),
            'ssh_username':  self.user.text(),
            'ssh_password':  self.pwd.text(),
            'enable_secret': self.sec.text()
        }
        save_cfg(cfg)

        self.model.removeRows(0,self.model.rowCount())
        self.pbar.setValue(0)
        self.start_btn.setEnabled(False)

        self.thread = QThread(self)
        self.worker = Worker(
            self.excel_in.text(),
            self.out_in.text(),
            self.user.text(),
            self.pwd.text(),
            self.sec.text(),
            selected_sites
        )
        self.worker.moveToThread(self.thread)
        self.worker.device_done.connect(self.on_device)
        self.worker.progress.connect(self.pbar.setValue)
        self.worker.finished.connect(self.on_finished)
        self.worker.finished.connect(self.thread.quit)
        self.worker.error.connect(lambda tb: QMessageBox.critical(self,"Error",tb))

        self.thread.started.connect(self.worker.run)
        self.thread.start()

    def on_device(self, idx, host, ip, ok, reason):
        r = self.model.rowCount()
        self.model.insertRow(r, [
            QStandardItem(str(idx+1)),
            QStandardItem(host),
            QStandardItem(ip),
            QStandardItem("✓" if ok else "✗"),
            QStandardItem(reason)
        ])

    def on_finished(self, out_path):
        QMessageBox.information(self, "Done", f"Report saved:\n{out_path}")
        self.start_btn.setEnabled(True)
        self.pbar.setValue(100)

if __name__=="__main__":
    app = QApplication([])
    app.setStyleSheet(qdarkstyle.load_stylesheet())
    win = MainWindow()
    win.show()
    app.exec()
