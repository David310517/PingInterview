# ========================
# 📁 SD-WAN Migration GUI Tool
# Developed by: David [YourLastName]
# Date: July 2025
# ========================

# ========== Imports ==========
import os
import re
import ipaddress
import datetime
import openpyxl
import configparser
import threading

from PySide6.QtCore import Signal, Qt
from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QFormLayout, QHBoxLayout,
    QVBoxLayout, QLabel, QLineEdit, QPushButton, QFileDialog, QComboBox,
    QSpinBox, QProgressBar, QMessageBox, QCheckBox
)
from PySide6.QtGui import QIcon
import qdarkstyle

# ========== Constants ==========
CFG_FILE = os.path.expanduser("~/.sdwan_migration_gui.ini")
PUBLIC_KWS = ("comcast", "crowncastle", "astound", "att")
PRIVATE_PATTERN = re.compile(r"\b\w+ans\d*\b", re.IGNORECASE)

# ========== Configuration Management ==========
def load_config() -> configparser.ConfigParser:
    cfg = configparser.ConfigParser()
    if os.path.exists(CFG_FILE):
        cfg.read(CFG_FILE)
    return cfg

def save_config(cfg: configparser.ConfigParser) -> None:
    with open(CFG_FILE, "w") as f:
        cfg.write(f)

# ========== Main GUI Window ==========
class MigrationWindow(QMainWindow):
    progress_signal = Signal(int)
    done_signal = Signal(str)

    def __init__(self):
        super().__init__()
        self.setWindowTitle("\U0001F310 SD‑WAN Migration Planner")
        self.setMinimumSize(800, 550)
        self.setWindowIcon(QIcon.fromTheme("network-workgroup"))

        self.cfg = load_config()
        defaults = self.cfg['DEFAULT'] if 'DEFAULT' in self.cfg else {}

        # === Load default settings ===
        process_all = defaults.getboolean('process_all', True)
        start_row = defaults.getint('start_row', 2)

        # === Layout setup ===
        central = QWidget()
        self.setCentralWidget(central)
        main_layout = QVBoxLayout(central)
        form = QFormLayout()
        main_layout.addLayout(form)

        # === Input widgets ===
        self.ci_input = QLineEdit(defaults.get('circuit_file', ''))
        btn_ci = QPushButton("Browse…")
        btn_ci.clicked.connect(self.browse_circuit)
        form.addRow("Circuit Info Workbook (.xlsx):", self.hbox(self.ci_input, btn_ci))

        self.mig_input = QLineEdit(defaults.get('migration_file', ''))
        btn_mig = QPushButton("Browse…")
        btn_mig.clicked.connect(self.browse_migration)
        form.addRow("Migration Planning Workbook (.xlsx):", self.hbox(self.mig_input, btn_mig))

        self.chk_all = QCheckBox("Process All Sheets")
        self.chk_all.setChecked(process_all)
        self.chk_all.stateChanged.connect(self.toggle_sheet_selection)
        form.addRow(self.chk_all)

        self.combo_sheet = QComboBox()
        self.combo_sheet.setEnabled(not process_all)
        form.addRow("Circuit Sheet (optional):", self.combo_sheet)

        self.spin_row = QSpinBox()
        self.spin_row.setRange(2, 10000)
        self.spin_row.setValue(start_row)
        form.addRow("Starting Row:", self.spin_row)

        self.out_input = QLineEdit(defaults.get('output_file', ''))
        btn_out = QPushButton("Browse…")
        btn_out.clicked.connect(self.browse_output)
        form.addRow("Output File (.xlsx):", self.hbox(self.out_input, btn_out))

        self.progress = QProgressBar()
        self.progress.setValue(0)
        main_layout.addWidget(self.progress)

        self.run_btn = QPushButton("Run Migration")
        self.run_btn.clicked.connect(self.start_migration)
        main_layout.addWidget(self.run_btn)

        # === Signal bindings ===
        self.progress_signal.connect(self.progress.setValue)
        self.done_signal.connect(self.show_done)

        if self.ci_input.text():
            self.load_sheets()

    # ========== Helper Layout Methods ==========
    def hbox(self, *widgets):
        box = QHBoxLayout()
        for w in widgets:
            box.addWidget(w)
        container = QWidget()
        container.setLayout(box)
        return container

    # ========== File Dialog Handlers ==========
    def browse_circuit(self):
        path, _ = QFileDialog.getOpenFileName(self, "Select Circuit Info", filter="Excel Files (*.xlsx)")
        if path:
            self.ci_input.setText(path)
            self.load_sheets()

    def browse_migration(self):
        path, _ = QFileDialog.getOpenFileName(self, "Select Migration Planning", filter="Excel Files (*.xlsx)")
        if path:
            self.mig_input.setText(path)

    def browse_output(self):
        default_name = self.out_input.text() or f"SD-WAN_Migration_{datetime.datetime.now():%Y%m%d_%H%M%S}.xlsx"
        path, _ = QFileDialog.getSaveFileName(self, "Select Output File", default_name, "Excel Files (*.xlsx)")
        if path:
            self.out_input.setText(path)

    def load_sheets(self):
        try:
            wb = openpyxl.load_workbook(self.ci_input.text(), read_only=True)
            self.combo_sheet.clear()
            self.combo_sheet.addItems(wb.sheetnames)
            self.toggle_sheet_selection()
        except Exception:
            self.combo_sheet.clear()
            self.combo_sheet.setEnabled(False)

    def toggle_sheet_selection(self):
        self.combo_sheet.setEnabled(not self.chk_all.isChecked())

    # ========== Status ==========
    def show_done(self, msg):
        QMessageBox.information(self, "Done", msg)
        self.run_btn.setEnabled(True)

    # ========== Run Button Logic ==========
    def start_migration(self):
        self.cfg['DEFAULT'] = {
            'circuit_file': self.ci_input.text(),
            'migration_file': self.mig_input.text(),
            'output_file': self.out_input.text(),
            'process_all': str(self.chk_all.isChecked()),
            'start_row': str(self.spin_row.value())
        }
        save_config(self.cfg)
        self.run_btn.setEnabled(False)
        threading.Thread(target=self.process).start()

    # ========== Core Processing Logic ==========
    def process(self):
        try:
            ci_wb = openpyxl.load_workbook(self.ci_input.text(), data_only=True)
            mig_wb = openpyxl.load_workbook(self.mig_input.text())
            target = "Transform Circuit Planning"
            ws = mig_wb[target] if target in mig_wb.sheetnames else mig_wb.active

            sheets = ci_wb.sheetnames if self.chk_all.isChecked() else ([self.combo_sheet.currentText()] if self.combo_sheet.currentText() else [])
            row = self.spin_row.value()
            total = len(sheets)

            for idx, sheet in enumerate(sheets, start=1):
                devices = parse_circuit_sheet(ci_wb[sheet])
                if devices:
                    ws.cell(row, 1).value = sheet
                    # === Primary circuit ===
                    p = devices[0]
                    ws.cell(row, 7).value = p['provider']
                    ws.cell(row, 8).value = p['raw_desc']
                    ws.cell(row, 9).value = p['handoff']
                    ws.cell(row, 10).value = p['bandwidth']
                    ws.cell(row, 11).value = p['vlan']
                    if p['ip'] and p['mask']:
                        ws.cell(row, 12).value = f"{p['ip']}/{prefix_len(p['mask'])}"
                    ws.cell(row, 13).value = p['peer']
                    # === Secondary circuit ===
                    if len(devices) > 1:
                        s = devices[1]
                        ws.cell(row, 16).value = s['provider']
                        ws.cell(row, 17).value = s['raw_desc']
                        ws.cell(row, 18).value = s['handoff']
                        ws.cell(row, 19).value = s['bandwidth']
                        ws.cell(row, 20).value = s['vlan']
                        if s['ip'] and s['mask']:
                            ws.cell(row, 21).value = f"{s['ip']}/{prefix_len(s['mask'])}"
                        ws.cell(row, 22).value = s['peer']
                    row += 1
                self.progress_signal.emit(int(idx / total * 100))

            mig_wb.save(self.out_input.text())
            self.done_signal.emit(f"Processed {len(sheets)} sheet(s) – output at {self.out_input.text()}")
        except Exception as e:
            self.done_signal.emit(f"Error: {e}")

# ========== Circuit Parsing Utilities ==========
def clean_desc(raw: str) -> str:
    d = raw.strip()
    d = re.sub(r"\s{2,}", " ", d)
    d = re.sub(r"\.{2,}", ".", d)
    d = re.sub(r"[\u2013\u2014-]", "–", d)
    return d[0].upper() + d[1:] if d else d

def prefix_len(mask: str) -> int:
    return ipaddress.IPv4Network(f"0.0.0.0/{mask}").prefixlen

def parse_circuit_sheet(ws):
    devices = []
    for col in ws.iter_cols(min_row=1, max_row=ws.max_row):
        entry = {k: "" for k in ['raw_desc','handoff','vlan','ip','mask','peer','bandwidth_line','bandwidth','provider','circuit_id']}
        for cell in col[1:]:
            line = str(cell.value or "").strip(); low = line.lower()
            if low.startswith("description "):
                entry['raw_desc'] = clean_desc(line[12:])
            elif low.startswith("bandwidth "):
                parts = line.split(); entry['bandwidth_line'] = parts[1] if len(parts) > 1 else ""
            elif low.startswith("ip address "):
                p = line.split(); entry['ip'], entry['mask'] = p[2], p[3]
                try:
                    net = ipaddress.ip_network(f"{entry['ip']}/{entry['mask']}", strict=False)
                    hosts = list(net.hosts())
                    entry['peer'] = hosts[1].compressed if hosts and hosts[0].compressed == entry['ip'] else hosts[0].compressed
                except: pass
            if "fiber" in low: entry['handoff'] = 'Fiber'
            elif "copper" in low: entry['handoff'] = 'Copper'
            if low.startswith("interface ") and '.' in line.split()[1]:
                entry['vlan'] = line.split()[1].split('.',1)[1]
        rd, rd_l = entry['raw_desc'], entry['raw_desc'].lower()
        pub_k = next((kw for kw in PUBLIC_KWS if kw in rd_l), ""); pub = pub_k.title() if pub_k else ""
        pm = PRIVATE_PATTERN.search(rd_l); priv = pm.group(0).lower() if pm else ""
        entry['provider'] = f"{pub},{priv}" if pub and priv else pub or priv
        cm = re.search(r"([A-Za-z0-9\.\-/]+)\s*$", rd); entry['circuit_id'] = cm.group(1) if cm else ""
        bm = re.search(r"(\d+)\s*(m|mb|meg|g|gb)(?![A-Za-z])", rd, re.IGNORECASE)
        if bm:
            n, u = bm.group(1), bm.group(2).lower(); entry['bandwidth'] = f"{n}{'G' if u.startswith('g') else 'M'}"
        else:
            bl = entry['bandwidth_line']; entry['bandwidth'] = bl if bl and bl.lower() != 'percent' else ''
        devices.append(entry)
    return devices

# ========== App Launch ==========
if __name__ == '__main__':
    app = QApplication([])
    app.setStyleSheet(qdarkstyle.load_stylesheet())
    wnd = MigrationWindow()
    wnd.show()
    app.exec()
